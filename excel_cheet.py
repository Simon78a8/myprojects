# -*- coding: utf-8 -*-
"""
Created on Mon Feb  7 13:13:51 2022

@author: simon
"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#wb = load_workbook('sampledatafoodsales.xlsx')
#ws = wb.active
#print(ws['B2'].value)

#wb.save('sampledatafoodsales.xlsx')
#print(wb.sheetnames)
#print(wb['FoodSales'])

##  CREATE A NEWWORK SHEET wb.create_sheet('Test')

wb = load_workbook('simon.xlsx')
ws = wb.active
#ws.title = 'Data'

# ws.append(['Simon', 'is', 'Great', '!'])

for row in range(1, 10):
    for col in range(1, 4):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)
        


##merging_cless

ws.merge_cells('A1:D1')
ws.unmerge_cells('A1:D1')

##insert
ws.insert_rows(7)
ws.insert_rows(8)

ws.insert_cols(2)



